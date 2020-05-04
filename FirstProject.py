from configparser import ConfigParser
from selenium.webdriver import Chrome
import openpyxl
import time
from selenium.webdriver.support.select import Select


## fetching data from configuration file
objconfig=ConfigParser()
objconfig.read(".\\venv\\Config.cfg")
url = objconfig.get("Basics","url")
browser_exe_path = objconfig.get("Basics","path")
write_excel = objconfig.get("Basics","write_excel")
read_excel = objconfig.get("Basics","read_excel")

## Writing data to excel file
objwbook = openpyxl.Workbook()
wsheet = objwbook.create_sheet("FirstProject")

##Reading  data from excel sheet
objwbook2= openpyxl.load_workbook(read_excel)
wsheet2 = objwbook2["data"]
name = wsheet2.cell(1,1).value
email= wsheet2.cell(1,2).value
password = wsheet2.cell(1,3).value

try:
    ## created an object "driver" for chrome class
    driver= Chrome(executable_path=browser_exe_path)
    driver.get(url)
    driver.maximize_window()
    #print(driver.title)
    # #print(driver.current_url)
    # #print(driver.page_source)
    # ## fetching text from ui and asserting its value.
    text = driver.find_element_by_xpath("//label[@for='tab1']").text
    driver.find_element_by_xpath("//form[@name='register']/input[2]").send_keys(name)
    driver.find_element_by_xpath("//form[@name='register']/input[3]").send_keys(email)
    driver.find_element_by_xpath("//input[@name='fld_email']/following-sibling::input[1]").send_keys(password)
    #time.sleep(60)
    driver.find_element_by_name("fld_cpassword").send_keys(password)
    driver.find_element_by_id("datepicker").send_keys("05/04/2020")
    driver.find_element_by_name("phone").send_keys("9999937005")
    driver.find_element_by_xpath("//input[contains(@placeholder,'Add')]").send_keys("Test address")
    driver.find_element_by_xpath("//input[@value='office']").click()
    sex_dd = Select(driver.find_element_by_name("sex"))
    sex_dd.select_by_value("1")
    print(sex_dd.first_selected_option.text)
    county_dd = Select(driver.find_element_by_name("country"))
    county_dd.select_by_index(101)
    print(county_dd.first_selected_option.text)
    driver.find_element_by_name("terms").click()
    time.sleep(60)
    #driver.find_element_by_xpath("//input[@value='Sign up']").click()
    driver.quit()
    status = "Pass"
except:
    print("Element not found, please check the locator")
    status = "Fail"
finally:
    if (text == "Register"):
        wsheet.cell(1,1).value = "Text is Present"
    else:
        wsheet.cell(1,1).value = "Text is Not Present"
wsheet.cell(1,2).value = "Section heading is " + text
wsheet.cell(1,3).value = status
objwbook.save(write_excel)
