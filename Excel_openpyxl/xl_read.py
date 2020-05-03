import openpyxl

# object of workbook class
objwbook = openpyxl.load_workbook("C:\FileRead\openpyxl.xlsx")

# object for worksheet
#wsheet = objwbook.active
#print(wsheet.title)
sheet_names = objwbook.sheetnames
no_of_sheets = len(objwbook.worksheets)
# to select the particular sheet
wsheet1 = objwbook["data"]

print("Total no of worksheets are " + str(no_of_sheets))
print("Worksheet names are ", sheet_names)

print(wsheet1.title)

no_of_rows = wsheet1.max_row
no_of_colm = wsheet1.max_column
print(no_of_rows,no_of_colm)
# object for cell
wcell = wsheet1.cell(1,1)
print(wcell.value)

for i in sheet_names:
    wsheet = objwbook[i]
    for j in range(1, no_of_rows+1):
        for k in range(1, no_of_colm+1):
            if i=="data":
                print(wsheet.cell(j,k).value)
