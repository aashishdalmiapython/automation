import openpyxl

objwb= openpyxl.Workbook()
ws= objwb.active
ws.title = "Aashish Data"
ws.cell(1,1).value = "This is Pass"

#or we can create our own worksheet.
objwb.create_sheet("Dalmia Data")
ws1=objwb["Dalmia Data"]
ws1.cell(1,2).value ="This is Fail"


objwb.save("C:\\FileRead\\openpyxlwrite.xlsx")
